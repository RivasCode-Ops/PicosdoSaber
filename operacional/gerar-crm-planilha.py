#!/usr/bin/env python3
"""Gera o template Excel do CRM operacional Picos do Saber."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = Path(__file__).parent / "CRM-Template.xlsx"

STATUS_LABELS = [
    "\U0001f7e1 Interessado",
    "\U0001f535 Contato realizado",
    "\U0001f7e0 Visita agendada",
    "\U0001f7e3 Avaliacao realizada",
    "\U0001f7e2 Matricula realizada",
    "\U0001f534 Nao interessado",
    "\u26ab Lista de espera",
]

# Mesma lista do formulario do site (index.html) e MODELO-DE-TURMAS.md
SERIES = [
    "4\u00ba ano fundamental",
    "5\u00ba ano fundamental",
    "6\u00ba ano fundamental",
    "7\u00ba ano fundamental",
    "8\u00ba ano fundamental",
    "9\u00ba ano fundamental",
    "1\u00ba ano ensino m\u00e9dio",
    "2\u00ba ano ensino m\u00e9dio",
    "3\u00ba ano ensino m\u00e9dio",
    "Refor\u00e7o para provas",
]

DIFICULDADES = [
    "Matem\u00e1tica",
    "Portugu\u00eas / leitura",
    "Ci\u00eancias",
    "Organiza\u00e7\u00e3o e estudos",
    "M\u00faltiplas disciplinas",
    "Outro",
]

HEADER_FILL = PatternFill("solid", fgColor="1B4332")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14, color="1B4332")


def style_header(ws, row=1):
    for cell in ws[row]:
        if cell.value:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_col_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def build_dashboard(wb):
    ws = wb.active
    ws.title = "Dashboard"
    ws["A1"] = "PICOS DO SABER — CRM Template (copiar para planilha privada)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:C1")
    ws["A3"] = "Indicador"
    ws["B3"] = "Numero"
    ws["C3"] = "Observacao"
    style_header(ws, 3)

    rows = [
        ("Leads recebidos", "=COUNTA('CRM Leads'!A2:A500)", "Total de contatos registrados"),
        ("Contatos realizados", "=COUNTIF('CRM Leads'!I2:I500,\"*Contato realizado*\")", "Status: Contato realizado"),
        ("Visitas agendadas", "=COUNTIF('CRM Leads'!I2:I500,\"*Visita agendada*\")", "Status: Visita agendada"),
        ("Avaliacoes realizadas", "=COUNTIF('CRM Leads'!I2:I500,\"*Avaliacao realizada*\")", "Status: Avaliacao realizada"),
        ("Matriculas fechadas", "=COUNTIF('CRM Leads'!I2:I500,\"*Matricula realizada*\")", "Status: Matricula realizada"),
        ("Lista de espera (CRM)", "=COUNTIF('CRM Leads'!I2:I500,\"*Lista de espera*\")", "Leads aguardando vaga"),
        ("Lista de espera (aba)", "=COUNTA('Lista de Espera'!A2:A500)", "Fila oficial de espera"),
        ("Vagas totais", "=SUM('Turmas e Vagas'!C2:C50)", "Soma da capacidade"),
        ("Vagas disponiveis", "=SUM('Turmas e Vagas'!F2:F50)", "Soma das vagas livres"),
    ]

    start = 4
    for i, (label, formula, note) in enumerate(rows):
        r = start + i
        ws.cell(r, 1, label)
        ws.cell(r, 2, formula)
        ws.cell(r, 3, note)

    ws["A14"] = "Atualizado em:"
    ws["B14"] = "=TODAY()"
    ws["B14"].number_format = "DD/MM/YYYY"
    set_col_widths(ws, [28, 14, 42])


def build_crm(wb):
    ws = wb.create_sheet("CRM Leads")
    headers = [
        "Data",
        "Responsavel",
        "WhatsApp",
        "Aluno",
        "Serie",
        "Dificuldade",
        "Turno",
        "Origem",
        "Status",
        "Proxima acao",
        "Data retorno",
        "Observacoes",
    ]
    ws.append(headers)
    style_header(ws)

    ws.append([
        "DD/MM/AAAA",
        "EXEMPLO — apagar antes de usar",
        "(89) 90000-0000",
        "EXEMPLO — apagar",
        "6\u00ba ano fundamental",
        "Matem\u00e1tica",
        "Tarde",
        "Site",
        STATUS_LABELS[0],
        "Proxima acao",
        "DD/MM/AAAA",
        "Linha ficticia do template",
    ])

    status_list = ",".join(STATUS_LABELS)
    dv_status = DataValidation(
        type="list",
        formula1=f'"{status_list}"',
        allow_blank=True,
    )
    dv_status.error = "Escolha um status da lista padronizada."
    dv_status.prompt = "Status do funil comercial"
    ws.add_data_validation(dv_status)
    dv_status.add("I2:I500")

    for col, options in {
        "E": ",".join(SERIES),
        "F": ",".join(DIFICULDADES),
        "G": "Manha,Tarde,Noite",
        "H": "Site,Indicacao,Redes sociais,Panfleto,WhatsApp direto,Outro",
    }.items():
        dv = DataValidation(type="list", formula1=f'"{options}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}500")

    set_col_widths(ws, [12, 18, 16, 16, 12, 14, 10, 14, 22, 16, 14, 28])
    ws.freeze_panes = "A2"


def build_turmas(wb):
    ws = wb.create_sheet("Turmas e Vagas")
    headers = ["Turma", "Turno", "Capacidade", "Matriculados", "Pre-reservas", "Vagas livres", "Situacao"]
    ws.append(headers)
    style_header(ws)

    samples = [
        ("6o ano", "Manha", 8, 0, 0),
        ("7o ano", "Manha", 8, 0, 0),
        ("4o ao 5o ano", "Tarde", 8, 0, 0),
        ("8o ao 9o ano", "Tarde", 8, 0, 0),
        ("Ensino Medio", "Noite", 8, 0, 0),
        ("Reforco para provas", "Noite", 8, 0, 0),
    ]
    for turma, turno, cap, mat, pre in samples:
        ws.append([turma, turno, cap, mat, pre, None, None])

    for row in range(2, 2 + len(samples)):
        ws[f"F{row}"] = f"=C{row}-D{row}-E{row}"
        ws[f"G{row}"] = f'=IF(F{row}<=0,"Lista espera",IF(F{row}<=2,"Quase lotada","Aberta"))'

    note = ws.cell(10, 1, "Regra: vaga ocupada = matricula confirmada. Pre-reserva tem prazo. Turma lotada = lista de espera.")
    note.font = Font(italic=True, color="555555")
    ws.merge_cells("A10:G10")
    set_col_widths(ws, [14, 10, 12, 14, 14, 14, 16])
    ws.freeze_panes = "A2"


def build_matriculas(wb):
    ws = wb.create_sheet("Matriculas")
    headers = [
        "Aluno",
        "Responsavel",
        "Serie",
        "Turma",
        "Inicio",
        "Mensalidade",
        "Vencimento",
        "Situacao",
    ]
    ws.append(headers)
    style_header(ws)
    ws.append([
        "EXEMPLO — apagar",
        "EXEMPLO — apagar",
        "6o ano",
        "6o Manha",
        "DD/MM/AAAA",
        0,
        "Dia 10",
        "Ativo",
    ])

    dv = DataValidation(type="list", formula1='"Ativo,Inativo,Trancado"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("H2:H500")

    set_col_widths(ws, [16, 18, 12, 14, 12, 14, 12, 12])
    ws.freeze_panes = "A2"


def build_lista_espera(wb):
    ws = wb.create_sheet("Lista de Espera")
    headers = ["Ordem", "Data", "Aluno", "Serie", "Turno", "Contato", "Observacao"]
    ws.append(headers)
    style_header(ws)
    ws.append([
        1,
        "DD/MM/AAAA",
        "EXEMPLO — apagar",
        "7o ano",
        "Manha",
        "(89) 90000-0000",
        "Linha ficticia do template",
    ])

    for row in range(2, 52):
        ws[f"A{row}"] = f'=IF(C{row}<>"",COUNTA($C$2:C{row}),"")'

    set_col_widths(ws, [8, 12, 16, 12, 10, 18, 32])
    ws.freeze_panes = "A2"


def main():
    wb = Workbook()
    build_dashboard(wb)
    build_crm(wb)
    build_turmas(wb)
    build_matriculas(wb)
    build_lista_espera(wb)
    wb.save(OUTPUT)
    print(f"Planilha gerada: {OUTPUT}")


if __name__ == "__main__":
    main()
